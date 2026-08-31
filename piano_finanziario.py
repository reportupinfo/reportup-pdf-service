"""Piano Finanziario editabile — Excel allegato al Report Strategico.

Costruito dagli stessi campi deterministici già calcolati per il PDF da
_arricchisci_report_deterministico + _calcola_valore_asset in app.py (stesso
immobile, stessi numeri di partenza). Il cliente può poi cambiare le proprie
ipotesi (prezzo, occupazione, costi) e vedere il risultato ricalcolato in
Excel con le stesse formule del motore.

Due blocchi separati: GESTIONE (host — quanto rende gestendolo) e
INVESTIMENTO (investitore — quanto vale come asset). Solo le celle numeriche
di input restano editabili; le altre (label, risultati, formule) sono
bloccate e con formula nascosta nella barra formule (Protection hidden=True
+ foglio protetto da password) — vedi conversazione 31/8/2026 sul motivo
della scelta: il cliente deve poter usare il proprio file, non rivendere il
motore di calcolo copiandone le formule.
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation

PWD = "reportup2026"  # password unica prodotto Strategico — protegge il foglio, non il file

INPUT_FILL = PatternFill("solid", fgColor="FFF6D6")
LOCK_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FILL = PatternFill("solid", fgColor="1F3B57")      # blocco GESTIONE
HEADER_FILL_B = PatternFill("solid", fgColor="6E5A1F")    # blocco INVESTIMENTO
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
SUBHEADER_FONT = Font(color="FFFFFF", italic=True, size=8)
LABEL_FONT = Font(size=10)
RESULT_FONT = Font(size=10, bold=True, color="1F3B57")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LOCKED = Protection(locked=True, hidden=True)
UNLOCKED = Protection(locked=False, hidden=False)


def _section_header(ws, row, text, span=3, fill=None, sub=None):
    fill = fill or HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.protection = LOCKED
    for col in range(2, span + 1):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).protection = LOCKED
    ws.row_dimensions[row].height = 22
    if sub:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        sc = ws.cell(row=row, column=1, value=sub)
        sc.fill = fill
        sc.font = SUBHEADER_FONT
        sc.protection = LOCKED
        for col in range(2, span + 1):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).protection = LOCKED
        ws.row_dimensions[row].height = 16
    return row


def _input_row(ws, row, label, value, unit=""):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = LABEL_FONT; lc.protection = LOCKED; lc.border = BORDER
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = LABEL_FONT
    vc.fill = INPUT_FILL
    vc.protection = UNLOCKED
    vc.border = BORDER
    vc.number_format = "#,##0" if unit != "%" else "0"
    uc = ws.cell(row=row, column=3, value=unit)
    uc.font = LABEL_FONT; uc.protection = LOCKED; uc.border = BORDER
    return vc.coordinate


def _result_row(ws, row, label, formula, fmt="#,##0 €"):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = LABEL_FONT; lc.protection = LOCKED; lc.border = BORDER
    vc = ws.cell(row=row, column=2, value=formula)
    vc.font = RESULT_FONT
    vc.fill = LOCK_FILL
    vc.protection = LOCKED
    vc.border = BORDER
    vc.number_format = fmt
    return vc.coordinate


LEGENDA_VOCI = [
    ("BLOCCO GESTIONE — cosa modifichi", "", ""),
    ("Prezzo medio a notte", "Prezzo medio realizzato a notte occupata, dal tuo report ReportUp.",
     "Cambialo se pensi di posizionarti diversamente."),
    ("Occupazione media annua", "% di notti dell'anno effettivamente prenotate su quelle disponibili.",
     "64% = occupato 64 notti su 100."),
    ("Commissioni piattaforma", "Trattenuta annua di Airbnb/Booking ecc. sulle prenotazioni.",
     "Metti l'importo reale se lo conosci dai tuoi estratti."),
    ("Pulizie", "Costo annuo del servizio pulizie tra un ospite e l'altro.", ""),
    ("Biancheria", "Costo annuo lavaggio/sostituzione lenzuola e asciugamani.", ""),
    ("Utenze", "Luce, gas, acqua, internet, condominio a carico della gestione B&B.", ""),
    ("Manutenzione e ristrutturazione", "Stima libera: quanto pensi di spendere ogni anno per manutenere o rinnovare l'immobile.",
     "Nessuna formula dietro: è una tua previsione."),
    ("Gestione esterna / property manager", "Compenso annuo se affidi la gestione a terzi.",
     "Lascia 0 se gestisci tu in prima persona."),
    ("Mutuo attivo / Rata mutuo mensile", "SI/NO se hai un mutuo sull'immobile, e importo mensile se attivo.",
     "La rata non cambia con l'occupazione: si paga uguale a casa piena o vuota."),
    ("Bonus prenotazioni dirette", "Ricavo extra da prenotazioni fuori piattaforma (sito proprio, passaparola).",
     "5-10% in più sul ricavo da notti è la forbice tipica, 7% il valore medio."),
    ("", "", ""),
    ("BLOCCO GESTIONE — risultati calcolati", "", ""),
    ("Notti occupate/anno", "Quante notti l'anno il tuo alloggio è prenotato.", "= 365 × occupazione %"),
    ("Ricavi da pernottamenti", "Incasso dalle sole notti vendute.", "= prezzo/notte × notti occupate"),
    ("Ricavi totali annui", "Ricavi da pernottamenti + bonus prenotazioni dirette.", ""),
    ("Costi variabili", "Somma di commissioni, pulizie, biancheria, utenze, manutenzione, gestione esterna.", ""),
    ("Costi totali annui", "Costi variabili + costo mutuo annuo (se presente).", ""),
    ("Profitto netto annuo", "Ricavi totali − costi totali.", "Quanto ti resta in tasca ogni anno gestendo l'attività."),
    ("Margine", "Percentuale di ogni euro incassato che resta di profitto.", "= profitto netto / ricavi totali"),
    ("", "", ""),
    ("BLOCCO INVESTIMENTO — cosa modifichi", "", ""),
    ("Saggio di capitalizzazione", "Rendimento annuo che un investitore si aspetta da un asset di questo tipo.",
     "Più alto = più rischio percepito = valore stimato più basso. 7% è lo standard usato nel tuo report per B&B in Italia."),
    ("", "", ""),
    ("BLOCCO INVESTIMENTO — risultati calcolati", "", ""),
    ("EBITDA stimato", "Profitto prima degli oneri finanziari (il mutuo viene riaggiunto al profitto netto).",
     "Misura la redditività dell'attività a prescindere da come hai finanziato l'acquisto — è il numero che un investitore usa per confrontare asset diversi tra loro."),
    ("Valore stimato come asset B&B", "EBITDA capitalizzato al saggio scelto sopra.",
     "Quanto varrebbe oggi il tuo B&B come attività avviata, non come semplice immobile vuoto."),
]


def _build_legenda(wb):
    ws = wb.create_sheet("Legenda")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 40

    r = 1
    ws.merge_cells(f"A{r}:C{r}")
    t = ws.cell(row=r, column=1, value="Legenda e istruzioni — Piano Finanziario ReportUp")
    t.font = Font(bold=True, size=14, color="1F3B57")
    t.protection = LOCKED
    r += 1
    ws.merge_cells(f"A{r}:C{r}")
    intro = ws.cell(
        row=r, column=1,
        value=("Compila SOLO le celle gialle nel foglio \"Piano Finanziario\". Tutto il resto è "
               "protetto: contiene i calcoli e non va toccato per avere risultati corretti. "
               "I tuoi dati restano solo tuoi, il file è tuo."),
    )
    intro.font = Font(italic=True, size=10, color="444444")
    intro.alignment = Alignment(wrap_text=True, vertical="top")
    intro.protection = LOCKED
    ws.row_dimensions[r].height = 40
    r += 2

    for i, h in enumerate(["Voce", "Cosa significa", "Come si usa"]):
        c = ws.cell(row=r, column=1 + i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.protection = LOCKED
    ws.row_dimensions[r].height = 20
    r += 1

    for voce, significato, uso in LEGENDA_VOCI:
        if voce and not significato and not uso:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            c = ws.cell(row=r, column=1, value=voce)
            c.font = Font(bold=True, size=10, color="6E5A1F" if "INVESTIMENTO" in voce else "1F3B57")
            c.protection = LOCKED
            r += 1
            continue
        if not voce:
            r += 1
            continue
        for col, val in enumerate((voce, significato, uso)):
            c = ws.cell(row=r, column=1 + col, value=val)
            c.font = Font(size=9, bold=(col == 0))
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.protection = LOCKED
            c.border = BORDER
        r += 1

    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.set_password(PWD)


def build_piano_finanziario_bytes(data, cliente="", indirizzo="", ordine="", data_str=""):
    """data: il dict deterministico già arricchito per il PDF Strategico
    (dopo _arricchisci_report_deterministico + _calcola_valore_asset).
    Ritorna i byte di un .xlsx pronto da allegare alla mail."""
    prezzo_notte = data.get("prezzo_notte_stimato") or 0
    occupazione_percent = data.get("occupazione_percent") or 0
    commissioni_annue = data.get("costi_commissioni") or 0
    pulizie_annue = data.get("costi_pulizie") or 0
    biancheria_annua = data.get("costi_biancheria") or 0
    utenze_annue = data.get("costi_utenze") or 0
    manutenzione_annua = data.get("costi_manutenzione") or 0
    gestione_esterna_annua = 0  # non modellata dal motore: il cliente la aggiunge se la ha
    mutuo_attivo = "SI" if data.get("mutuo_attivo") else "NO"
    rata_mutuo_mensile = data.get("rata_mutuo_mensile") or 0
    bonus_dirette_percent = 7  # fisso, stesso valore usato dal motore deterministico (bonus = 7% ricavo lordo)
    saggio_capitalizzazione = data.get("saggio_capitalizzazione") or 7.0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Piano Finanziario"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10

    r = 1
    ws.merge_cells(f"A{r}:C{r}")
    t = ws.cell(row=r, column=1, value="ReportUp — Piano Finanziario Personalizzato")
    t.font = Font(bold=True, size=14, color="1F3B57")
    t.protection = LOCKED
    r += 1
    wm = ws.cell(row=r, column=1,
                 value=f"{cliente} — {indirizzo} — ordine {ordine} — {data_str}".strip(" —"))
    wm.font = Font(italic=True, size=8, color="999999")
    wm.protection = LOCKED
    ws.merge_cells(f"A{r}:C{r}")
    r += 2

    r = _section_header(ws, r, "GESTIONE — quanto puoi guadagnare gestendolo",
                         sub="Cambia solo le celle gialle. Vedi il foglio Legenda per il significato di ogni voce.")
    r += 1
    r_prezzo = _input_row(ws, r, "Prezzo medio a notte", prezzo_notte, "€"); r += 1
    r_occ = _input_row(ws, r, "Occupazione media annua", occupazione_percent, "%"); r += 1
    r_comm = _input_row(ws, r, "Commissioni piattaforma (annue)", commissioni_annue, "€"); r += 1
    r_puliz = _input_row(ws, r, "Pulizie (annue)", pulizie_annue, "€"); r += 1
    r_bianch = _input_row(ws, r, "Biancheria (annua)", biancheria_annua, "€"); r += 1
    r_utenze = _input_row(ws, r, "Utenze (annue)", utenze_annue, "€"); r += 1
    r_manut = _input_row(ws, r, "Manutenzione e ristrutturazione (annua, stima tua)",
                          manutenzione_annua, "€"); r += 1
    r_gestione = _input_row(ws, r, "Gestione esterna / property manager (annua)",
                             gestione_esterna_annua, "€"); r += 1

    lc = ws.cell(row=r, column=1, value="Mutuo attivo (SI/NO)")
    lc.font = LABEL_FONT; lc.protection = LOCKED; lc.border = BORDER
    mc = ws.cell(row=r, column=2, value=mutuo_attivo)
    mc.font = LABEL_FONT; mc.fill = INPUT_FILL; mc.protection = UNLOCKED; mc.border = BORDER
    dv = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(mc)
    r += 1
    r_rata = _input_row(ws, r, "Rata mutuo mensile", rata_mutuo_mensile, "€"); r += 1
    r_bonus = _input_row(ws, r, "Bonus prenotazioni dirette (fuori piattaforma)",
                          bonus_dirette_percent, "%"); r += 1

    r += 1
    r_notti = _result_row(ws, r, "Notti occupate/anno", f"=ROUND(365*{r_occ}/100,0)", "#,##0"); r += 1
    r_ricavi_base = _result_row(ws, r, "Ricavi da pernottamenti", f"={r_prezzo}*{r_notti}"); r += 1
    r_ricavi_tot = _result_row(ws, r, "Ricavi totali annui", f"={r_ricavi_base}*(1+{r_bonus}/100)"); r += 1
    r_costi_var = _result_row(ws, r, "Costi variabili (comm.+pulizie+biancheria+utenze+manut.+gestione)",
                               f"={r_comm}+{r_puliz}+{r_bianch}+{r_utenze}+{r_manut}+{r_gestione}"); r += 1
    r_mutuo_annuo = _result_row(ws, r, "Costo mutuo annuo",
                                 f'=IF({mc.coordinate}="SI",{r_rata}*12,0)'); r += 1
    r_costi_tot = _result_row(ws, r, "Costi totali annui", f"={r_costi_var}+{r_mutuo_annuo}"); r += 1
    r_profitto = _result_row(ws, r, "Profitto netto annuo", f"={r_ricavi_tot}-{r_costi_tot}"); r += 1
    r_margine = _result_row(ws, r, "Margine", f"=IF({r_ricavi_tot}=0,0,{r_profitto}/{r_ricavi_tot}*100)", "0.0%"); r += 1

    r += 1
    r = _section_header(ws, r, "INVESTIMENTO — quanto vale come asset", fill=HEADER_FILL_B,
                         sub="Solo per chi valuta il B&B come bene da comprare/vendere, non da gestire.")
    r += 1
    r_saggio = _input_row(ws, r, "Saggio di capitalizzazione", saggio_capitalizzazione, "%"); r += 1
    r_ebitda = _result_row(ws, r, "EBITDA stimato", f"={r_profitto}+{r_mutuo_annuo}"); r += 1
    r_valore = _result_row(ws, r, "Valore stimato come asset B&B",
                            f"=IF({r_saggio}=0,0,{r_ebitda}/{r_saggio}*100)"); r += 1

    r += 1
    note = ws.cell(row=r, column=1,
                    value="Cambia solo le celle gialle. Le altre sono protette per garantire calcoli corretti.")
    note.font = Font(italic=True, size=8, color="999999")
    note.protection = LOCKED
    ws.merge_cells(f"A{r}:C{r}")

    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False
    ws.protection.deleteRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.set_password(PWD)

    _build_legenda(wb)

    wb.security = openpyxl.workbook.protection.WorkbookProtection(
        workbookPassword=None, lockStructure=True
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
